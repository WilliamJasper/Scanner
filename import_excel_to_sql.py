"""
Import Excel to SQL Server (ExcelTtbDB).
Usage: python import_excel_to_sql.py <excel_path> [table_name]

Table name defaults to OCR_TTB_IMPORT.
Connection uses DATABASE_URL from .env or falls back to local SQL Server.
"""
import os
import re
import sys
from pathlib import Path

import openpyxl
import pyodbc
from dotenv import load_dotenv

load_dotenv()


def get_connection_params() -> dict:
    """Parse DATABASE_URL to get pyodbc connection params."""
    url = os.getenv("DATABASE_URL", "").strip()
    server = "."
    driver = "ODBC Driver 17 for SQL Server"
    trusted = True
    if url and "mssql+pyodbc://" in url:
        ms = re.search(r"mssql\+pyodbc://([^/]+)/", url)
        if ms:
            server = ms.group(1).strip()
        md = re.search(r"[?&]driver=([^&]+)", url)
        if md:
            driver = md.group(1).replace("+", " ").strip()
        mt = re.search(r"[?&]trusted_connection=([^&]+)", url, re.I)
        if mt:
            trusted = str(mt.group(1)).lower() in ("yes", "true", "1")
    return {"server": server, "driver": driver, "trusted": trusted}


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python import_excel_to_sql.py <excel_path> [table_name]")

    excel_path = Path(sys.argv[1])
    table_name = sys.argv[2] if len(sys.argv) > 2 else "OCR_TTB_IMPORT"
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    raw_headers = [c.value for c in ws[1]]
    headers = []
    seen = {}
    for i, h in enumerate(raw_headers, start=1):
        name = str(h).strip() if h is not None else ""
        if not name:
            name = f"Column_{i}"
        name = re.sub(r"\s+", " ", name)
        base = name
        n = seen.get(base, 0) + 1
        seen[base] = n
        if n > 1:
            name = f"{base}_{n}"
        headers.append(name)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append([None if v is None else str(v) for v in row])

    params = get_connection_params()
    base_conn = (
        f"Driver={{{params['driver']}}};Server={params['server']};"
        + ("Trusted_Connection=yes;" if params["trusted"] else "")
        + "TrustServerCertificate=yes;"
    )

    conn = pyodbc.connect(base_conn + "Database=master;", autocommit=True)
    cur = conn.cursor()
    cur.execute("IF DB_ID('ExcelTtbDB') IS NULL CREATE DATABASE [ExcelTtbDB];")
    conn.close()

    conn = pyodbc.connect(base_conn + "Database=ExcelTtbDB;")
    cur = conn.cursor()
    safe_table = re.sub(r"[^A-Za-z0-9_]", "_", table_name.strip())[:120] or "OCR_TTB_IMPORT"
    cur.execute(f"IF OBJECT_ID('dbo.{safe_table}','U') IS NOT NULL DROP TABLE dbo.{safe_table};")

    col_defs = ", ".join([f"[{h.replace(']', ']]')}] NVARCHAR(MAX) NULL" for h in headers])
    cur.execute(f"CREATE TABLE dbo.{safe_table} ({col_defs});")

    placeholders = ", ".join(["?"] * len(headers))
    col_list = ", ".join([f"[{h.replace(']', ']]')}]" for h in headers])
    insert_sql = f"INSERT INTO dbo.{safe_table} ({col_list}) VALUES ({placeholders})"
    cur.fast_executemany = True
    cur.executemany(insert_sql, rows)
    conn.commit()

    cur.execute(f"SELECT COUNT(*) FROM dbo.{safe_table};")
    row_count = cur.fetchone()[0]
    cur.execute(
        f"SELECT COUNT(*) FROM sys.columns WHERE object_id = OBJECT_ID('dbo.{safe_table}');"
    )
    col_count = cur.fetchone()[0]

    print(f"DB=ExcelTtbDB TABLE=dbo.{safe_table} ROWS={row_count} COLS={col_count}")
    conn.close()


if __name__ == "__main__":
    main()
